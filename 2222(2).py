import os
import jsonpath
import openpyxl
import requests

exl = 'test_case_api.xlsx'


def excel_master_data(excel_name, sheet_name):
    if os.path.exists(excel_name):
        test_excel = openpyxl.load_workbook(excel_name)
        sheet = test_excel[sheet_name]
        list1 = []
        max_row = sheet.max_row
        for i in range(2, max_row + 1, 1):
            dict1 = dict(
                case_id=sheet.cell(row=i, column=1).value,
                interface=sheet.cell(row=i, column=2).value,
                method=sheet.cell(row=i, column=4).value,
                url=sheet.cell(row=i, column=5).value,
                data=sheet.cell(row=i, column=6).value,
                expected=sheet.cell(row=i, column=7).value
            )
            list1.append(dict1)
        return list1
    else:
        return '文件不存在'


def update_excel_expected(exl_name, exl_sheet_name, value, value2, update_name):
    test_excle = openpyxl.load_workbook(exl_name)
    sheet = test_excle[exl_sheet_name]
    result = sheet.cell(row=value, column=value2)
    result.value = update_name
    test_excle.save(exl_name)
    return update_name


def public_res(method, body, url, head={"X-Lemonban-Media-Type": "lemonban.v2",
                                        "Content-Type": "application/json"}):
    if method.lower() == 'post':
        res_value = requests.post(url=url, json=body, headers=head).json()
    if method.lower() == 'patch':
        res_value = requests.patch(url=url, json=body, headers=head).json()
    if method.lower() == 'get':
        res_value = requests.get(url=url, json=body, headers=head).json()
    return res_value


def automatic(excel_1, sheet, log=None):
    res_excel = excel_master_data(excel_1, sheet)
    a = 0
    if not isinstance(res_excel, str):
        a += 1
        list_1 = []
        for excel in res_excel:
            register_expected = excel['expected']
            if excel['interface'] == 'register' or excel['interface'] == 'login':
                res = public_res(excel['method'], eval(excel['data']), excel['url'])
            else:
                login_res = public_res(method='post',
                                       url='http://api.lemonban.com/futureloan/member/login',
                                       body=log)
                dl_id = jsonpath.jsonpath(login_res, '$..id')[0]
                token = jsonpath.jsonpath(login_res, '$..token')[0]
                token_head = {"X-Lemonban-Media-Type": "lemonban.v2",
                              "Content-Type": "application/json",
                              "Authorization": "Bearer" + " " + token
                              }
                if excel['interface'] == 'loan_add':
                    cz_data = eval(excel['data'])
                    cz_data['member_id'] = dl_id
                    res = public_res(excel['method'], cz_data, excel['url'], head=token_head)
                else:
                    if excel['data'] is not None:
                        res = public_res(excel['method'], eval(excel['data']), excel['url'],
                                         head=token_head)
                    else:
                        if excel['url'] == 'http://120.78.128.25:8766/futureloan/member/{}/info':
                            get_lur = excel['url'].format(dl_id)
                            res = public_res(excel['method'], None, get_lur, head=token_head)
                        else:
                            res = public_res(excel['method'], None, excel['url'], head=token_head)
            print('-' * 30)
            print('case_id: {}'.format(excel['case_id']))
            print('预期结果{}'.format(eval(register_expected)))
            res_expected = {'code': res['code'], 'msg': res['msg']}
            print('实际结果{}'.format(res_expected))
            if eval(register_expected) == res_expected:
                print('通过')
                cg = update_excel_expected(excel_1, sheet, excel['case_id'] + 1, 8, '通过')
                consequence = {'sheet': sheet, 'case_id': excel['case_id'], 'result': cg}
                list_1.append(consequence)
            else:
                print('不通过')
                sb = update_excel_expected(excel_1, sheet, excel['case_id'] + 1, 8, '不通过')
                consequence = {'sheet': sheet, 'case_id': excel['case_id'], 'result': sb}
                list_1.append(consequence)
        return list_1
    else:
        return '文件不存在'


# 除了注册登录，其他操作都需先登录。请设置登录手机号和密码
# 普通账号
login_body = {
    "mobile_phone": "15815541555",
    "pwd": "lemon123456"
}
# 管理员账号
login0_body = {
    "mobile_phone": "15815541666",
    "pwd": "lemon123456"
}
# 注册-----------------------------------------
print(automatic(exl, 'register'))
# 登录-----------------------------------------
print(automatic(exl, 'login'))
# 充值-----------------------------------------
# print(automatic(exl, 'recharge', login_body))
# 加标-----------------------------------------
# print(automatic(exl, 'loan_add', login_body))
# 审核-----------------------------------------
# print(automatic(exl, 'loan_audit', login0_body))
# 获取用户信息-----------------------------------------
# print(automatic(exl, 'user_info', login0_body))
# 获取项目列表-----------------------------------------
# print(automatic(exl, 'loans', login0_body))
